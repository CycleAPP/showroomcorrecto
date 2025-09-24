#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Sube TODAS las imágenes incrustadas en un .xlsx a Cloudinary (no incremental).
- Toma el nombre del public_id desde la columna '2026 model' (con aliases).
- Usa EXCEL_URL (Google Sheets export) o EXCEL_PATH.
- SHEET_NAME='Master' y HEADER_ROW=5 por defecto.
- Python 3.9 compatible (sin typing PEP604).
- Sin dependencia de 'requests'; sólo urllib.

Requiere: openpyxl, cloudinary   (opcional: python-dotenv, Pillow para convertir EMF/WMF/BMP a PNG)
"""

import os, re, io, sys, json, zipfile, time
from urllib.request import urlopen, Request
from xml.etree import ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
import cloudinary, cloudinary.uploader

# ---- .env opcional ----
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

# ---- Opcional: convertir formatos raros a PNG (EMF/WMF/BMP) ----
try:
    from PIL import Image
    PIL_OK = True
except Exception:
    PIL_OK = False

def maybe_convert_to_png(img_bytes):
    if not PIL_OK:
        return img_bytes
    try:
        im = Image.open(io.BytesIO(img_bytes))
        out = io.BytesIO()
        im.save(out, format="PNG", optimize=True)
        return out.getvalue()
    except Exception:
        return img_bytes

# ========= CONFIG =========
EXCEL_PATH   = (os.getenv("EXCEL_PATH") or "").strip()
EXCEL_URL    = (os.getenv("EXCEL_URL")  or "").strip()
SHEET_NAME   = (os.getenv("SHEET_NAME") or "Master").strip()
HEADER_ROW   = int((os.getenv("HEADER_ROW") or "5").strip())

# Columna modelo (aliases)
A_MODEL = [s.strip() for s in (os.getenv("COL_MODEL") or "2026 model,2026 Model,Model,Item #,Item,Modelo,#Item,Item#").split(",")]

# Cloudinary
CLOUD_NAME = (os.getenv("CLOUDINARY_CLOUD_NAME") or "").strip()
API_KEY    = (os.getenv("CLOUDINARY_API_KEY") or "").strip()
API_SECRET = (os.getenv("CLOUDINARY_API_SECRET") or "").strip()
CLD_FOLDER = (os.getenv("CLOUDINARY_FOLDER") or "showroom_2025").strip()

# Performance
CLD_CONCURRENCY  = int(os.getenv("CLD_CONCURRENCY") or "8")      # hilos para subir
CLD_TIMEOUT      = int(os.getenv("CLD_TIMEOUT") or "120")        # s (timeout por upload en Cloudinary)
CLD_MAX_RETRIES  = int(os.getenv("CLD_MAX_RETRIES") or "2")      # reintentos por imagen

# Logs
VERBOSE = True

# Salida informativa
DATA_DIR       = "data"
IMG_MAP_JSON   = os.path.join(DATA_DIR, "cloudinary_map_full.json")
os.makedirs(DATA_DIR, exist_ok=True)

# ========= VALIDACIONES =========
if not (EXCEL_PATH or EXCEL_URL):
    sys.exit("❌ ERROR: Proporciona EXCEL_PATH o EXCEL_URL.")
if not (CLOUD_NAME and API_KEY and API_SECRET):
    sys.exit("❌ ERROR: Falta CLOUDINARY_CLOUD_NAME / CLOUDINARY_API_KEY / CLOUDINARY_API_SECRET.")

# ========= Cloudinary init =========
cloudinary.config(cloud_name=CLOUD_NAME, api_key=API_KEY, api_secret=API_SECRET, secure=True, timeout=CLD_TIMEOUT)

# ========= Utils =========
def norm(s): return re.sub(r"\s+", " ", (s or "")).strip().lower()

def sanitize_filename(s):
    s = (s or "").strip()
    s = re.sub(r"[^\w\-]+", "_", s)
    s = s.strip("_")
    return s or "unnamed"

def fuzzy_find_key(row_dict, aliases):
    keys = list(row_dict.keys())
    nmap = {norm(k): k for k in keys}
    for a in aliases:
        k = nmap.get(norm(a))
        if k: return k
    # contains
    for k in keys:
        nk = norm(k)
        for a in aliases:
            if norm(a) in nk:
                return k
    return None

def fetch_excel_bytes():
    if EXCEL_PATH:
        with open(EXCEL_PATH, "rb") as f:
            return f.read()
    req = Request(EXCEL_URL, headers={"User-Agent":"Mozilla/5.0"})
    with urlopen(req, timeout=120) as r:
        return r.read()

def logln(msg):
    if VERBOSE:
        print(msg, flush=True)

# ========= Cargar libro/hoja =========
start_total = time.time()
logln("[init] Descargando/leyendo Excel…")
xlsx_bytes = fetch_excel_bytes()
wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
if SHEET_NAME not in wb.sheetnames:
    logln("⚠️  Hoja '%s' no existe, usando la primera: %s" % (SHEET_NAME, wb.sheetnames[0]))
    SHEET_NAME = wb.sheetnames[0]
ws = wb[SHEET_NAME]

headers = [cell.value if cell.value is not None else "" for cell in ws[HEADER_ROW]]

def row_to_dict(r):
    row = ws[r]
    d = {}
    for i, cell in enumerate(row[:len(headers)]):
        d[headers[i]] = cell.value if cell.value is not None else ""
    return d

sample_dict = row_to_dict(HEADER_ROW + 1)
model_key = fuzzy_find_key(sample_dict, A_MODEL)
if not model_key:
    print("Headers:", headers)
    sys.exit("❌ ERROR: No pude detectar la columna del modelo. Aliases: %s" % A_MODEL)

# ========= Parsear drawings del .xlsx =========
NS = {
    "a":   "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r":   "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
}

z = zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r")

sheet_index = wb.sheetnames.index(ws.title) + 1
sheet_rels_path = "xl/worksheets/_rels/sheet%d.xml.rels" % sheet_index

drawing_target = None
if sheet_rels_path in z.namelist():
    rels_xml = ET.fromstring(z.read(sheet_rels_path))
    RNS = {"": "http://schemas.openxmlformats.org/package/2006/relationships"}
    for rel in rels_xml.findall("Relationship", RNS):
        if rel.attrib.get("Type","").endswith("/drawing"):
            drawing_target = rel.attrib.get("Target")
            break

if not drawing_target:
    with open(IMG_MAP_JSON, "w", encoding="utf-8") as f:
        json.dump({}, f, ensure_ascii=False, indent=2)
    sys.exit("ℹ️ La hoja no tiene imágenes incrustadas.")

drawing_path = drawing_target.replace("../", "xl/")
if not drawing_path.startswith("xl/"):
    drawing_path = "xl/" + drawing_target.lstrip("/")

if drawing_path not in z.namelist():
    sys.exit("❌ No encontré el drawing en el ZIP: %s" % drawing_path)

rels_path = drawing_path.replace("drawings/", "drawings/_rels/") + ".rels"
rels_map = {}
if rels_path in z.namelist():
    d_rels = ET.fromstring(z.read(rels_path))
    RNS = {"": "http://schemas.openxmlformats.org/package/2006/relationships"}
    for rel in d_rels.findall("Relationship", RNS):
        rId = rel.attrib.get("Id")
        tgt = rel.attrib.get("Target")
        media_path = tgt.replace("../", "xl/")
        if not media_path.startswith("xl/"):
            media_path = "xl/" + tgt.lstrip("/")
        rels_map[rId] = media_path

drawing_xml = ET.fromstring(z.read(drawing_path))
anchors = []
for tag in ("twoCellAnchor", "oneCellAnchor"):
    for anc in drawing_xml.findall("xdr:%s" % tag, NS):
        blip = anc.find(".//a:blip", NS)
        if blip is None:
            continue
        rId = blip.attrib.get("{%s}embed" % NS["r"])
        media = rels_map.get(rId)
        if not media:
            continue
        frm = anc.find("xdr:from", NS)
        if frm is None:
            continue
        row_el = frm.find("xdr:row", NS)
        col_el = frm.find("xdr:col", NS)
        if row_el is None or col_el is None:
            continue
        row0 = int(row_el.text or "0")
        col0 = int(col_el.text or "0")
        anchors.append({"row0": row0, "col0": col0, "media": media})

if not anchors:
    with open(IMG_MAP_JSON, "w", encoding="utf-8") as f:
        json.dump({}, f, ensure_ascii=False, indent=2)
    sys.exit("ℹ️ No se detectaron anchors de imágenes.")

logln("[ok] Anchors detectados: %d" % len(anchors))

def find_model_for_row(start_r):
    row_excel = start_r
    if row_excel <= HEADER_ROW:
        row_excel = HEADER_ROW + 1
    if row_excel > ws.max_row:
        return "", None
    rdict = row_to_dict(row_excel)
    val = (rdict.get(model_key) or "").strip() if rdict else ""
    if val:
        return val, row_excel
    for off in range(1, 4):
        r2 = row_excel + off
        if r2 > ws.max_row: break
        rd2 = row_to_dict(r2)
        mv2 = (rd2.get(model_key) or "").strip() if rd2 else ""
        if mv2:
            return mv2, r2
    return "", row_excel

# ========= Construir lista (modelo, bytes) =========
items = []  # [(model, img_bytes)]
for a in anchors:
    row_excel = a["row0"] + 1
    model_val, _ = find_model_for_row(row_excel)
    if not model_val:
        continue

    media_path = a["media"]
    if media_path not in z.namelist():
        # fallback por número
        mnum = re.search(r"image(\d+)\.(\w+)", media_path or "", re.I)
        chosen = None
        if mnum:
            for c in z.namelist():
                if c.startswith("xl/media/") and ("image%s." % mnum.group(1)) in c:
                    chosen = c; break
        media_path = chosen or media_path

    if media_path not in z.namelist():
        logln("⚠️  Media no encontrada en ZIP: %s" % media_path)
        continue

    img_bytes = z.read(media_path)
    # Conversión opcional para formatos pesados/raros
    if re.search(r"\.(emf|wmf|bmp)$", media_path, re.I):
        img_bytes = maybe_convert_to_png(img_bytes)

    items.append((model_val, img_bytes))

if not items:
    with open(IMG_MAP_JSON, "w", encoding="utf-8") as f:
        json.dump({}, f, ensure_ascii=False, indent=2)
    sys.exit("ℹ️ No se hallaron imágenes utilizable.")

# ========= Subida paralela =========
uploaded = 0
failed   = 0
model_to_url = {}

def upload_with_retry(img_bytes, public_id):
    last_err = None
    for attempt in range(1, CLD_MAX_RETRIES+1):
        try:
            return cloudinary.uploader.upload(
                img_bytes,
                folder=CLD_FOLDER,
                public_id=public_id,
                overwrite=True,           # no incremental: siempre sobrescribe
                unique_filename=False,
                resource_type="image",
                timeout=CLD_TIMEOUT
            )
        except Exception as e:
            last_err = e
            time.sleep(min(2*attempt, 6))  # backoff corto
    raise last_err

def worker(idx, total_n, model, img_bytes):
    public_id = sanitize_filename(model)
    logln("[prog] %d/%d → %s" % (idx, total_n, public_id))
    res = upload_with_retry(img_bytes, public_id)
    secure = res.get("secure_url", "")
    optimized = secure.replace("/upload/", "/upload/f_auto,q_auto/")
    return (model, optimized)

logln("[run] Subiendo %d imágenes a Cloudinary con %d hilos…" % (len(items), CLD_CONCURRENCY))
start_upload = time.time()

with ThreadPoolExecutor(max_workers=CLD_CONCURRENCY) as ex:
    futures = []
    total_n = len(items)
    for i, (model, bts) in enumerate(items, start=1):
        futures.append(ex.submit(worker, i, total_n, model, bts))

    # Heartbeat para que el front vea vida si tarda
    next_hb = time.time() + 10

    for fut in as_completed(futures):
        # heartbeats periódicos
        now = time.time()
        if now >= next_hb:
            logln("[hb] trabajando… %d/%d completados" % (uploaded+failed, len(items)))
            next_hb = now + 10

        try:
            model, url = fut.result()
            uploaded += 1
            model_to_url[model] = url
            logln("[ok] %s → %s" % (model, url))
        except Exception as e:
            failed += 1
            logln("[err] %s" % (e,))

elapsed_up = time.time() - start_upload
logln("[done] Uploads terminados en %.1fs" % elapsed_up)

# Guardar mapa (opcional)
with open(IMG_MAP_JSON, "w", encoding="utf-8") as f:
    json.dump(model_to_url, f, ensure_ascii=False, indent=2)

elapsed_total = time.time() - start_total
print("\n========== RESUMEN ==========")
print("Anchors detectados:   ", len(anchors))
print("Intentos de subida:   ", len(items))
print("Subidas exitosas:     ", uploaded)
print("Fallidas:             ", failed)
print("Carpeta Cloudinary:   ", CLD_FOLDER)
print("Mapa (debug):         ", IMG_MAP_JSON)
print("Tiempo total:         ", "%.1fs" % elapsed_total)
print("================================")
